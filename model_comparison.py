"""
model_comparison.py
Modul evaluasi model AI untuk Sawit Vision.

Membandingkan centroid hasil inference (satu atau beberapa model) terhadap
centroid manual (ground truth) memakai greedy nearest-neighbor matching
(mirip evaluasi objek deteksi: 1 titik manual hanya boleh dipasangkan
dengan 1 titik inference terdekat dalam radius threshold).

Tidak bergantung pada QGIS. Hanya butuh: pyshp, scipy, numpy, openpyxl
(semua ringan, tidak ada dependensi GDAL/geopandas).
"""
import os
import json
import sqlite3
import struct

import numpy as np
import shapefile  # pyshp
from scipy.spatial import cKDTree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUPPORTED_EXTS = (".shp", ".gpkg", ".geojson", ".json")


def _geom_centroid_and_bbox(geom):
    """Hitung centroid + axis-aligned bounding box dari geometri shapefile.
    Return: dict {"centroid": (x,y), "bbox": (xmin,ymin,xmax,ymax)}, atau
            None kalau geometri kosong.

    Untuk POINT: bbox = None (jelas gak punya box).
    Untuk POLYGON: bbox = envelope dari semua vertex, centroid = rata-rata
                   vertex unik.
    """
    pts = geom.points
    if not pts:
        return None
    # Untuk polygon pyshp, vertex pertama = vertex terakhir (ring tertutup).
    if len(pts) >= 2 and pts[0] == pts[-1]:
        pts = pts[:-1]
    if len(pts) == 1:
        # Single point -- gak ada bbox konsep-nya
        return {"centroid": (float(pts[0][0]), float(pts[0][1])), "bbox": None}
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    cx = sum(xs) / len(xs)
    cy = sum(ys) / len(ys)
    bbox = (min(xs), min(ys), max(xs), max(ys))
    return {"centroid": (float(cx), float(cy)), "bbox": tuple(float(v) for v in bbox)}


def _geom_centroid(geom):
    """Wrapper kompatibilitas: cuma balikin centroid, tanpa bbox.
    Dipertahankan supaya code lain yang panggil ini gak ke-break."""
    result = _geom_centroid_and_bbox(geom)
    return result["centroid"] if result else None


def read_points_shp(path):
    """
    Baca semua "titik" dari shapefile. Untuk shapefile POLYGON (mis. hasil
    inference Sawit Vision yang menyimpan bounding box), centroid tiap polygon
    otomatis dihitung -- BUKAN ambil pojok pertama. Ini krusial supaya
    comparison tidak bias ~1 m ke arah pojok kiri-atas box.

    Return: (xy, attrs, bboxes)
      xy      -> np.ndarray shape (N, 2), centroid tiap objek
      attrs   -> list of dict atribut per titik (boleh kosong/tidak dipakai)
      bboxes  -> list of (xmin,ymin,xmax,ymax) atau None per objek.
                 None kalau geometri POINT (gak punya box).
                 List ini dipakai untuk containment-based matching:
                 matching berdasarkan "GT point masuk ke dalam bounding box
                 deteksi", bukan cuma jarak antar centroid.
    """
    sf = shapefile.Reader(path)
    field_names = [f[0] for f in sf.fields[1:]]  # field[0] = DeletionFlag, skip
    pts = []
    attrs = []
    bboxes = []
    for sr in sf.iterShapeRecords():
        result = _geom_centroid_and_bbox(sr.shape)
        if result is None:
            continue
        pts.append(result["centroid"])
        bboxes.append(result["bbox"])
        rec = list(sr.record)
        attrs.append(dict(zip(field_names, rec)))
    if not pts:
        return np.empty((0, 2), dtype=float), [], []
    return np.array(pts, dtype=float), attrs, bboxes


def _wkb_point_xy(wkb: bytes):
    """Ambil x,y dari WKB Point/PointZ/PointM/PointZM (2 titik pertama cukup buat centroid)."""
    order = "<" if wkb[0] == 1 else ">"
    x, y = struct.unpack(order + "dd", wkb[5:21])
    return x, y


def read_points_gpkg(path):
    """
    Baca titik dari GeoPackage (.gpkg) pakai sqlite3 bawaan Python (tanpa GDAL).
    Ambil layer geometri pertama yang terdaftar di gpkg_geometry_columns.
    Sekalian ambil kolom atribut lain di tabel yang sama (di luar kolom geometri)
    supaya bisa di-join dengan data model lain saat export perbandingan.

    Return: (xy, attrs, bboxes)
      bboxes -> selalu list of None untuk .gpkg (kita cuma baca POINT/POINTZ,
                gak baca polygon di sini karena GT manual = titik).
    """
    conn = sqlite3.connect(path)
    try:
        cur = conn.cursor()
        cur.execute("SELECT table_name, column_name FROM gpkg_geometry_columns")
        rows = cur.fetchall()
        if not rows:
            raise ValueError("Tidak ada layer geometri ditemukan di GeoPackage ini.")
        table, geom_col = rows[0]  # layer pertama; kalau ada beberapa layer, ambil yang pertama
        cur.execute(f'SELECT * FROM "{table}"')
        col_names = [d[0] for d in cur.description]
        attr_cols = [c for c in col_names if c != geom_col]
        geom_idx = col_names.index(geom_col)

        pts = []
        attrs = []
        for row in cur.fetchall():
            blob = row[geom_idx]
            if blob is None:
                continue
            # header GeoPackage: 'GP'(2) + version(1) + flags(1) + srs_id(4) [+ envelope]
            flags = blob[3]
            envelope_len = {0: 0, 1: 32, 2: 48, 3: 48, 4: 64}.get((flags >> 1) & 0x07, 0)
            wkb = blob[8 + envelope_len:]
            pts.append(_wkb_point_xy(wkb))
            attrs.append({c: row[col_names.index(c)] for c in attr_cols})
    finally:
        conn.close()
    if not pts:
        return np.empty((0, 2), dtype=float), [], []
    # bboxes semua None: GeoPackage di sini hanya support POINT (buat GT manual).
    return np.array(pts, dtype=float), attrs, [None] * len(pts)


def read_points_geojson(path):
    """Baca titik dari .geojson / .json (FeatureCollection Point/MultiPoint/Polygon->centroid).
    Kolom "properties" tiap feature ikut diambil sebagai atribut supaya bisa
    di-join dengan data model lain saat export perbandingan.

    Return: (xy, attrs, bboxes)
      bboxes -> untuk feature POLYGON, isi (xmin,ymin,xmax,ymax). Untuk POINT/
                MultiPoint, isi None."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    pts = []
    attrs = []
    bboxes = []
    for feat in data.get("features", []):
        geom = feat.get("geometry") or {}
        gtype = geom.get("type")
        coords = geom.get("coordinates")
        props = feat.get("properties") or {}
        if not coords:
            continue
        if gtype == "Point":
            pts.append((coords[0], coords[1]))
            attrs.append(dict(props))
            bboxes.append(None)
        elif gtype == "MultiPoint":
            for c in coords:
                pts.append((c[0], c[1]))
                attrs.append(dict(props))
                bboxes.append(None)
        elif gtype == "Polygon":
            ring = coords[0]
            xs = [c[0] for c in ring]
            ys = [c[1] for c in ring]
            pts.append((sum(xs) / len(xs), sum(ys) / len(ys)))
            attrs.append(dict(props))
            bboxes.append((min(xs), min(ys), max(xs), max(ys)))
    if not pts:
        return np.empty((0, 2), dtype=float), [], []
    return np.array(pts, dtype=float), attrs, bboxes


def read_points_any(path):
    """Dispatcher backward-compatible: baca titik dari .shp / .gpkg / .geojson.

    Return: (xy, attrs)  <-- SENGAJA hanya 2 elemen supaya code lama
                             yang unpack `xy, attrs = read_points_any(path)`
                             tetap jalan.

    Kalau lo butuh bounding box juga (untuk containment matching), pakai
    read_points_any_full() yang return 3 elemen (xy, attrs, bboxes).
    """
    xy, attrs, _bboxes = read_points_any_full(path)
    return xy, attrs


def read_points_any_full(path):
    """Dispatcher lengkap: return (xy, attrs, bboxes) untuk semua format.
    Pakai ini kalau lo butuh bounding box (untuk containment matching)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".shp":
        return read_points_shp(path)
    elif ext == ".gpkg":
        return read_points_gpkg(path)
    elif ext in (".geojson", ".json"):
        return read_points_geojson(path)
    raise ValueError(f"Format file tidak didukung: {ext} (pakai .shp / .gpkg / .geojson)")


def _match_greedy_single_group(manual_xy, infer_xy, threshold, manual_ids, infer_ids):
    """Greedy nearest-neighbor 1-to-1 untuk SATU kelompok (dalam 1 kelas, atau
    seluruh data kalau kelas tidak dipakai). `manual_ids`/`infer_ids` = index
    ASLI (global, sebelum di-subset per kelas) yang sejajar posisi dengan
    manual_xy/infer_xy -- dipetakan balik supaya hasil match/fp/fn tetap
    memakai index yang benar terhadap data lengkap."""
    n_manual, n_infer = len(manual_xy), len(infer_xy)
    if n_infer == 0:
        return [], [], list(manual_ids)
    if n_manual == 0:
        return [], list(infer_ids), []

    tree = cKDTree(manual_xy)
    dist, idx = tree.query(infer_xy, k=1)
    order = np.argsort(dist)  # proses dari yang paling dekat dulu

    used_manual = set()
    matches = []
    fp = []
    for i in order:
        d = float(dist[i])
        m = int(idx[i])
        if d <= threshold and m not in used_manual:
            used_manual.add(m)
            matches.append((infer_ids[i], manual_ids[m], d))
        else:
            fp.append(infer_ids[i])
    fn = [manual_ids[m] for m in range(n_manual) if m not in used_manual]
    return matches, fp, fn


def match_greedy(manual_xy: np.ndarray, infer_xy: np.ndarray, threshold: float,
                  manual_classes=None, infer_classes=None):
    """
    Greedy nearest-neighbor, one-to-one matching.
    Setiap titik manual maksimal dipasangkan 1x (mencegah 1 manual dihitung TP berkali-kali
    kalau ada beberapa deteksi tumpang tindih di sekitarnya).

    `manual_classes`/`infer_classes`: list label kelas per titik (opsional).
    Kalau KEDUANYA diberikan, pencocokan HANYA dilakukan antar titik kelas
    yang SAMA -- deteksi yang posisinya pas tapi kelasnya salah TIDAK akan
    dihitung TP lagi (sebelumnya matching ini murni jarak, jadi model 2 kelas
    yang salah tebak kelas tapi posisinya benar tetap dihitung benar).
    Titik yang kelasnya sama sekali tidak ada pasangannya di sisi lain
    otomatis FP/FN tanpa perlu dicek jaraknya. Kalau salah satu/keduanya
    None, fallback ke matching spasial murni (perilaku lama, backward
    compatible untuk model 1 kelas / GT tanpa atribut kelas).

    Return:
      matches -> list of (infer_idx, manual_idx, distance)
      fp_idx  -> list infer_idx yang TIDAK dapat pasangan (False Positive)
      fn_idx  -> list manual_idx yang TIDAK dapat pasangan (False Negative)
    """
    n_manual, n_infer = len(manual_xy), len(infer_xy)
    if n_infer == 0:
        return [], [], list(range(n_manual))
    if n_manual == 0:
        return [], list(range(n_infer)), []

    if manual_classes is None or infer_classes is None:
        return _match_greedy_single_group(
            manual_xy, infer_xy, threshold, list(range(n_manual)), list(range(n_infer)))

    all_classes = {c for c in manual_classes if c is not None} | \
        {c for c in infer_classes if c is not None}

    matches, fp, fn = [], [], []
    for c in all_classes:
        m_ids = [i for i in range(n_manual) if manual_classes[i] == c]
        i_ids = [i for i in range(n_infer) if infer_classes[i] == c]
        if not i_ids:
            fn.extend(m_ids)
            continue
        if not m_ids:
            fp.extend(i_ids)
            continue
        sub_matches, sub_fp, sub_fn = _match_greedy_single_group(
            manual_xy[m_ids], infer_xy[i_ids], threshold, m_ids, i_ids)
        matches.extend(sub_matches)
        fp.extend(sub_fp)
        fn.extend(sub_fn)
    return matches, fp, fn


def _extract_confidence(attr):
    """Cari nilai confidence dari dict atribut. Sawit Vision pakai field
    'confidence', tapi kita coba beberapa nama umum juga sebagai fallback
    (score/conf/probability) supaya robust ke depan.

    Return float 0..1 (atau angka apa saja), atau None kalau tidak ketemu.
    """
    if not attr:
        return None
    for key in ("confidence", "conf", "score", "prob", "probability"):
        if key in attr:
            try:
                v = attr[key]
                if v is None or v == "":
                    continue
                return float(v)
            except (TypeError, ValueError):
                continue
    return None


def _extract_class_label(attr):
    """Cari label kelas dari dict atribut. Coba beberapa nama field umum
    (shapefile Sawit Vision pakai 'kelas', tapi file GT manual bisa saja
    pakai nama lain) supaya robust tanpa perlu skema baku.

    Return string ternormalisasi (lower+strip) atau None kalau field kelas
    tidak ditemukan sama sekali -- None berarti "info kelas tidak tersedia",
    BUKAN "kelas kosong", supaya caller tahu harus fallback ke matching
    spasial murni (backward compatible untuk model 1 kelas / GT tanpa
    atribut kelas).
    """
    if not attr:
        return None
    for key in ("kelas", "class", "class_name", "kategori", "jenis", "label"):
        if key in attr:
            v = attr[key]
            if v is None or v == "":
                continue
            return str(v).strip().lower()
    return None


def match_containment(manual_xy, infer_xy, infer_bboxes, infer_attrs,
                       manual_classes=None, infer_classes=None):
    """Containment-based matching dengan confidence tiebreaker.

    Untuk setiap titik GT (manual), cari SEMUA bounding box deteksi yang
    MENGANDUNG titik itu (point-in-box check). Kalau ada ambiguitas
    (>1 box), pilih yang:
      1. Confidence tertinggi (primary tiebreaker)
      2. Jarak centroid box ke GT terkecil (secondary tiebreaker)
    Aturan 1-to-1: setiap box hanya boleh dipakai untuk 1 GT.

    `manual_classes`/`infer_classes`: list label kelas per titik (opsional).
    Kalau KEDUANYA diberikan, box yang secara SPASIAL mengandung titik GT
    tapi KELAS PREDIKSINYA BEDA dari kelas GT tidak dianggap kandidat valid
    sama sekali -- jadi tidak lagi bisa "menang" jadi TP hanya karena
    posisinya pas (sebelumnya matching ini murni containment+confidence,
    kelas prediksi tidak pernah dicek). Kalau salah satu/keduanya None,
    fallback ke containment murni (perilaku lama, backward compatible).

    Kenapa aturan 1-to-1: mencegah model "curang" -- 1 box gede yang
    mencakup 5 pohon tidak boleh menghasilkan 5 TP. Kalau box sudah
    dipakai untuk GT_A, GT_B/C/D yang juga masuk box itu jadi FN.

    Return:
      matches -> list of (infer_idx, manual_idx, distance)
                 distance di sini = jarak centroid box ke GT (untuk info,
                 bukan syarat matching).
      fp_idx  -> list infer_idx yang TIDAK dapat pasangan (deteksi yang
                 tidak mengandung GT manapun, atau kalah tiebreaker, atau
                 kelasnya tidak cocok dengan GT manapun yang ditampungnya).
      fn_idx  -> list manual_idx yang TIDAK dapat pasangan (GT yang tidak
                 masuk ke box manapun, atau hanya masuk ke box berkelas beda).
    """
    n_manual, n_infer = len(manual_xy), len(infer_xy)
    if n_infer == 0:
        return [], [], list(range(n_manual))
    if n_manual == 0:
        return [], list(range(n_infer)), []

    use_class = manual_classes is not None and infer_classes is not None

    # Ekstrak confidence tiap deteksi untuk tiebreaker. Kalau None,
    # anggap 0.0 (bakal kalah tiebreaker vs yang punya confidence).
    confidences = [
        _extract_confidence(attr) if attr else None
        for attr in infer_attrs
    ]

    # Untuk tiap GT, cari semua box yang mengandungnya + hitung jarak
    # centroid box ke GT (untuk secondary tiebreaker).
    #
    # Kita bangun list of (gt_idx, list_of_candidates) di mana
    # candidates = [(infer_idx, conf, dist_to_gt), ...]
    # yang dipilih adalah candidate dengan (conf tertinggi, dist terkecil).

    used_infer = set()  # box yang sudah dipakai (aturan 1-to-1)
    matches = []
    fn = []

    # Iterasi tiap GT
    for m_idx, (gx, gy) in enumerate(manual_xy):
        gt_class = manual_classes[m_idx] if use_class else None
        candidates = []
        for i_idx, bbox in enumerate(infer_bboxes):
            if i_idx in used_infer:
                continue  # sudah dipakai GT lain
            if bbox is None:
                continue  # box POINT, gak bisa untuk containment
            if use_class and infer_classes[i_idx] != gt_class:
                continue  # box menampung titik GT tapi KELAS BEDA -- bukan match valid
            xmin, ymin, xmax, ymax = bbox
            if xmin <= gx <= xmax and ymin <= gy <= ymax:
                cx, cy = infer_xy[i_idx]
                dist = float(np.hypot(cx - gx, cy - gy))
                conf = confidences[i_idx] if confidences[i_idx] is not None else 0.0
                candidates.append((i_idx, conf, dist))

        if not candidates:
            fn.append(m_idx)
            continue

        # Pilih candidate terbaik: conf DESC, dist ASC.
        # sorted key: (-conf, dist) supaya sorted() menaruh yang terbaik di index 0.
        candidates.sort(key=lambda c: (-c[1], c[2]))
        best_i, _best_conf, best_dist = candidates[0]
        used_infer.add(best_i)
        matches.append((best_i, m_idx, best_dist))

    # Deteksi yang tidak dipakai = FP
    fp = [i for i in range(n_infer) if i not in used_infer]
    return matches, fp, fn


def evaluate_model(manual_xy, infer_xy, threshold, infer_bboxes=None, infer_attrs=None,
                    manual_attrs=None):
    """Hitung metrik lengkap untuk satu model.

    MODE MATCHING:
    - Kalau infer_bboxes tersedia (list dengan minimal 1 bbox tidak-None) →
      pakai CONTAINMENT matching. Threshold TIDAK dipakai (containment ini
      bounded oleh ukuran box itu sendiri). Tiebreaker: confidence, lalu
      jarak centroid.
    - Kalau infer_bboxes semua None atau tidak diberikan → fallback ke
      GREEDY DISTANCE matching (perilaku lama). Threshold dipakai sebagai
      radius maksimum matching.

    Auto-detect ini bikin API backward-compatible: kalau caller lama
    manggil `evaluate_model(manual_xy, infer_xy, threshold)` tanpa bbox,
    tetap jalan seperti biasa. Kalau caller baru pass bbox, otomatis
    dapat containment matching yang lebih tepat.

    CLASS-AWARE MATCHING: kalau `manual_attrs` DAN `infer_attrs` sama-sama
    diberikan dan sama-sama punya field kelas yang terbaca (lihat
    _extract_class_label -- coba "kelas"/"class"/dst), matching hanya
    dianggap valid antar titik BERKELAS SAMA. Deteksi yang posisinya pas
    tapi salah tebak kelas tidak lagi dihitung TP. Ini otomatis nonaktif
    (fallback ke matching spasial murni, perilaku lama) kalau salah satu
    sisi tidak punya info kelas sama sekali -- jadi tetap backward
    compatible untuk model 1 kelas / GT tanpa atribut kelas.
    """
    manual_classes = [_extract_class_label(a) for a in manual_attrs] if manual_attrs else None
    infer_classes = [_extract_class_label(a) for a in infer_attrs] if infer_attrs else None
    class_aware = bool(
        manual_classes is not None and infer_classes is not None and
        any(c is not None for c in manual_classes) and
        any(c is not None for c in infer_classes)
    )
    if not class_aware:
        manual_classes = None
        infer_classes = None

    # Auto-detect mode
    has_bbox = (infer_bboxes is not None and
                any(b is not None for b in infer_bboxes))

    if has_bbox:
        matches, fp, fn = match_containment(
            manual_xy, infer_xy, infer_bboxes, infer_attrs or [{}] * len(infer_xy),
            manual_classes=manual_classes, infer_classes=infer_classes,
        )
        match_mode = "containment"
    else:
        matches, fp, fn = match_greedy(
            manual_xy, infer_xy, threshold,
            manual_classes=manual_classes, infer_classes=infer_classes,
        )
        match_mode = "distance"

    tp = len(matches)
    n_fp = len(fp)
    n_fn = len(fn)
    precision = tp / (tp + n_fp) if (tp + n_fp) > 0 else 0.0
    recall = tp / (tp + n_fn) if (tp + n_fn) > 0 else 0.0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0
    dists = np.array([d for _, _, d in matches]) if matches else np.array([])
    metrics = {
        "n_manual": len(manual_xy),
        "n_infer": len(infer_xy),
        "tp": tp, "fp": n_fp, "fn": n_fn,
        "precision": precision, "recall": recall, "f1": f1,
        "mean_dist": float(dists.mean()) if dists.size else 0.0,
        "median_dist": float(np.median(dists)) if dists.size else 0.0,
        "rmse_dist": float(np.sqrt((dists ** 2).mean())) if dists.size else 0.0,
        "max_dist": float(dists.max()) if dists.size else 0.0,
        "match_mode": match_mode,  # 'containment' atau 'distance' -- info buat log
        "class_aware": class_aware,  # True kalau kelas ikut dicek saat matching
    }
    return metrics, matches, fp, fn


# ============================================================
# AUTO-COMPUTE THRESHOLD dari struktur spasial GT
# ============================================================
# Prinsip: threshold yang tepat untuk matching centroid TIDAK boleh
# ditebak dari asumsi literatur -- harus turun dari data GT itu sendiri.
# Rumus: threshold = (median_nearest_neighbor_distance / 2) * safety_factor.
# Kenapa median/2: setengah jarak antar pohon = batas maksimum agar 1
# deteksi tidak "ambigu" antara 2 pohon GT tetangga.
# Kenapa safety 0.7: kasih margin agar tidak persis di ambang batas.

def compute_nn_stats(xy):
    """Hitung statistik nearest-neighbor distance dari array titik.
    Return dict dengan min/p10/p25/median/mean/std/p75/p90/max, atau None
    kalau jumlah titik < 2 (tidak cukup untuk analisis spasial)."""
    if xy is None or len(xy) < 2:
        return None
    tree = cKDTree(xy)
    # k=2 karena k=1 selalu return dirinya sendiri (jarak 0).
    dists, _ = tree.query(xy, k=2)
    nn = dists[:, 1]  # tetangga terdekat KEDUA = bukan diri sendiri
    return {
        "n_points": len(xy),
        "nn_min": float(nn.min()),
        "nn_p10": float(np.percentile(nn, 10)),
        "nn_p25": float(np.percentile(nn, 25)),
        "nn_median": float(np.median(nn)),
        "nn_mean": float(nn.mean()),
        "nn_std": float(nn.std()),
        "nn_p75": float(np.percentile(nn, 75)),
        "nn_p90": float(np.percentile(nn, 90)),
        "nn_max": float(nn.max()),
    }


def auto_compute_threshold(gt_path, safety_factor=0.7):
    """Hitung threshold matching yang direkomendasikan dari file GT.
    Fungsi ini yang dipanggil GUI saat user klik "Hitung Otomatis dari GT".

    Rumus:
        threshold = (median_nearest_neighbor / 2) * safety_factor

    Return dict:
      "threshold"     : float, threshold rekomendasi dalam satuan CRS GT
      "conservative"  : float, threshold ketat (median/4)
      "liberal"       : float, threshold longgar (median/2, tanpa safety)
      "nn_stats"      : dict statistik nearest-neighbor
      "explanation"   : string, penjelasan asal angka (buat di-log ke user)
      "n_points"      : int, jumlah titik GT
      "warning"       : string atau None, warning kalau ada anomali

    Raises ValueError kalau file tidak bisa dibaca atau titik < 2.
    """
    xy, _attrs = read_points_any(gt_path)
    n = len(xy)
    if n < 2:
        raise ValueError(
            f"GT hanya berisi {n} titik. Butuh minimal 2 titik untuk hitung "
            "jarak antar-tetangga. Threshold otomatis tidak bisa dihitung."
        )

    stats = compute_nn_stats(xy)
    median_nn = stats["nn_median"]
    p25_nn = stats["nn_p25"]

    # Skenario aneh: median NN sangat kecil (< 0.5m) -- kemungkinan
    # ada titik duplikat di GT. Fallback ke p25 (lebih robust) dan kasih warning.
    warning = None
    if median_nn < 0.5:
        base_spacing = p25_nn if p25_nn >= 0.5 else 1.0
        warning = (f"Median jarak antar-tetangga di GT sangat kecil "
                   f"({median_nn:.2f} m). Kemungkinan ada titik duplikat. "
                   f"Threshold dihitung dari fallback (p25 atau 1.0 m).")
    else:
        base_spacing = median_nn

    threshold = round(base_spacing / 2.0 * safety_factor, 2)
    conservative = round(base_spacing / 4.0, 2)
    liberal = round(base_spacing / 2.0, 2)

    explanation = (
        f"Auto-threshold: berdasarkan {n} titik GT. "
        f"Median jarak antar-tetangga = {median_nn:.2f} m. "
        f"Threshold = ({base_spacing:.2f} / 2) x {safety_factor} = {threshold} m. "
        f"Range: konservatif {conservative} m .. longgar {liberal} m."
    )

    return {
        "threshold": threshold,
        "conservative": conservative,
        "liberal": liberal,
        "nn_stats": stats,
        "explanation": explanation,
        "n_points": n,
        "warning": warning,
    }


def _numeric_avg(values):
    """Rata-rata dari sekumpulan nilai, lewati yang bukan angka / kosong."""
    nums = []
    for v in values:
        if v is None or v == "":
            continue
        try:
            nums.append(float(v))
        except (TypeError, ValueError):
            continue
    return float(np.mean(nums)) if nums else None


def _numeric_attr_keys(attrs_list):
    """Cari nama kolom atribut yang isinya angka (untuk dihitung rata-ratanya)."""
    keys = []
    seen = set()
    for a in attrs_list:
        for k, v in a.items():
            if k in seen:
                continue
            seen.add(k)
            try:
                if v is not None and v != "":
                    float(v)
                    keys.append(k)
            except (TypeError, ValueError):
                continue
    return keys


def _safe_sheet_name(name: str, used: set) -> str:
    base = "".join(c for c in name if c not in '[]:*?/\\')[:26] or "Model"
    candidate = f"Detail_{base}"[:31]
    n = 1
    while candidate in used:
        n += 1
        candidate = f"Detail_{base}_{n}"[:31]
    used.add(candidate)
    return candidate


# ============================================================
# STYLING HELPERS -- dipakai bersama oleh semua sheet biar tampilannya
# konsisten dan lebih enak dibaca (rapi, tidak berantakan).
# ============================================================
_COLOR_HEADER_FILL = "2FBF71"     # hijau brand Sawit Vision
_COLOR_HEADER_TEXT = "FFFFFF"
_COLOR_BORDER = "D9D9D9"
_COLOR_BAND = "F2F9F6"            # selang-seling baris (banding) supaya mudah diikuti mata
_COLOR_BEST = "C6EFCE"            # highlight model terbaik (F1 tertinggi)
_COLOR_TP = "E2F0D9"              # hijau muda -> deteksi benar
_COLOR_FP = "FCE4E4"              # merah muda -> deteksi salah/duplikat
_COLOR_FN = "FFF2CC"              # kuning muda -> pohon manual yang tidak ketemu

_THIN = Side(style="thin", color=_COLOR_BORDER)
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header_row(ws, row_idx: int = 1, n_cols: int = None):
    n_cols = n_cols or ws.max_column
    ws.row_dimensions[row_idx].height = 22
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = Font(bold=True, color=_COLOR_HEADER_TEXT)
        cell.fill = PatternFill("solid", fgColor=_COLOR_HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER_ALL


def _apply_borders(ws, min_row=1, max_row=None, min_col=1, max_col=None):
    max_row = max_row or ws.max_row
    max_col = max_col or ws.max_column
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _BORDER_ALL


def _apply_row_banding(ws, header_row: int, last_row: int, n_cols: int, skip_fill=None):
    """Kasih warna selang-seling ke baris data (bukan header) supaya baris demi baris
    lebih gampang diikuti mata saat scroll data yang panjang. skip_fill: set baris
    (nomor absolut) yang TIDAK boleh ditimpa banding (mis. baris yang sudah di-highlight)."""
    skip_fill = skip_fill or set()
    for r in range(header_row + 1, last_row + 1):
        if r in skip_fill:
            continue
        if (r - header_row) % 2 == 0:
            fill = PatternFill("solid", fgColor=_COLOR_BAND)
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = fill


def _autofit_columns(ws, min_width=10, max_width=42, padding=2, start_row=1):
    widths = {}
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            col = cell.column_letter
            widths[col] = max(widths.get(col, 0), length)
    for col, length in widths.items():
        ws.column_dimensions[col].width = max(min_width, min(max_width, length + padding))


def _apply_number_formats(ws, headers, start_row, start_col=1):
    """Format kolom berdasarkan nama header-nya: persen untuk Precision/Recall/F1,
    3 desimal untuk kolom jarak (m), integer untuk kolom hitungan (N/TP/FP/FN),
    4 desimal untuk kolom rata-rata atribut (Avg ...)."""
    for offset, header in enumerate(headers):
        col_idx = start_col + offset
        if header in ("Precision", "Recall", "F1"):
            fmt = "0.0%"
        elif header.endswith("(m)"):
            fmt = "0.000"
        elif header in ("N Manual", "N Deteksi", "TP", "FP", "FN"):
            fmt = "0"
        elif header.startswith("Avg "):
            fmt = "0.0000"
        elif header == "distance_m":
            fmt = "0.000"
        elif header.endswith("_x") or header.endswith("_y") or header == "confidence":
            fmt = "0.0000"
        else:
            continue
        for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value != "" and cell.value is not None:
                    cell.number_format = fmt


def _write_glossary_sheet(wb, threshold: float):
    """Sheet 'Keterangan': penjelasan istilah & rumus, supaya orang yang baru lihat
    Excel ini (mis. dosen pembimbing / pihak yang menilai) tidak perlu tanya-tanya
    arti tiap kolom."""
    ws = wb.create_sheet("Keterangan")
    ws.append(["Istilah / Kolom", "Penjelasan"])
    _style_header_row(ws, n_cols=2)

    rows = [
        ("N Manual", "Jumlah titik centroid manual (ground truth) yang jadi acuan penilaian."),
        ("N Deteksi", "Jumlah total titik hasil deteksi model AI (sebelum dicocokkan ke titik manual)."),
        ("TP (True Positive)",
         "Deteksi model yang berhasil dipasangkan dengan 1 titik manual dalam radius "
         f"threshold ({threshold} m). Dianggap deteksi yang BENAR."),
        ("FP (False Positive)",
         "Deteksi model yang TIDAK berhasil dipasangkan dengan titik manual manapun "
         "(dalam radius threshold). Bisa berarti deteksi palsu, objek bukan sawit, "
         "atau duplikat yang belum ter-suppress."),
        ("FN (False Negative)",
         "Titik manual yang TIDAK berhasil ditemukan oleh model manapun. Artinya pohon "
         "yang seharusnya ada tapi terlewat/tidak terdeteksi."),
        ("Precision", "Rumus: TP / (TP + FP). Dari SEMUA yang dideteksi model, berapa persen "
                       "yang benar-benar sawit (bukan deteksi palsu/duplikat)."),
        ("Recall", "Rumus: TP / (TP + FN). Dari SEMUA pohon yang sebenarnya ada (manual), "
                    "berapa persen yang berhasil ditemukan model."),
        ("F1", "Rumus: 2 x (Precision x Recall) / (Precision + Recall). Skor tunggal yang "
               "menyeimbangkan Precision & Recall -- dipakai sebagai acuan utama membandingkan model."),
        ("Mean Dist (m)", "Rata-rata jarak (meter) antara titik TP dan pasangan manual-nya. "
                           "Makin kecil, makin presisi posisi deteksi model."),
        ("Median Dist (m)", "Nilai tengah dari semua jarak TP. Lebih tahan terhadap outlier "
                             "dibanding Mean Dist."),
        ("RMSE Dist (m)", "Root Mean Square Error jarak TP. Lebih sensitif terhadap jarak yang "
                           "besar (outlier) dibanding Mean Dist -- kalau RMSE jauh lebih besar "
                           "dari Mean, artinya ada beberapa pasangan yang jaraknya jauh meleset."),
        ("Max Dist (m)", "Jarak TP terjauh -- kasus pencocokan terburuk pada model ini."),
        ("Threshold jarak", f"Radius maksimal ({threshold} m) supaya sebuah deteksi dianggap "
                             "\"cocok\" (match) dengan titik manual terdekatnya."),
        ("Avg Infer.<kolom>", "Rata-rata nilai atribut numerik (mis. confidence) dari SEMUA "
                               "deteksi model ini, bukan cuma yang TP."),
        ("Avg Manual.<kolom> (TP)", "Rata-rata nilai atribut numerik dari titik manual, dihitung "
                                     "HANYA dari pasangan yang berhasil match (TP)."),
        ("Catatan \u2605 F1 Tertinggi", "Model dengan skor F1 paling tinggi di antara yang "
                                         "dibandingkan -- kandidat model terbaik berdasarkan "
                                         "keseimbangan Precision & Recall."),
        ("status (sheet Detail)", "TP = deteksi benar (hijau), FP = deteksi salah/duplikat "
                                   "(merah), FN = pohon manual terlewat (kuning)."),
    ]
    for term, desc in rows:
        ws.append([term, desc])

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 30

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 95
    ws.freeze_panes = "A2"
    _apply_borders(ws)
    return ws


def export_comparison_excel(output_path, manual_xy, model_results, threshold, manual_attrs=None):
    """
    model_results: list of dict:
        {"name": str, "xy": np.ndarray, "metrics": dict, "matches": list, "fp": list, "fn": list,
         "attrs": list of dict (atribut per titik inference, boleh kosong)}
    manual_attrs: list of dict, atribut per titik centroid manual (boleh None/kosong).

    Struktur workbook:
      1. Ringkasan   -- tabel utama perbandingan model (dengan highlight F1 tertinggi)
      2. Keterangan  -- penjelasan istilah & rumus tiap kolom (TP, Precision, dst)
      3. Info        -- metadata run (threshold, jumlah model, dsb)
      4. Detail_<model> -- baris per titik (TP/FP/FN) dengan warna status
    """
    manual_attrs = manual_attrs or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkasan"

    base_headers = ["Model", "N Manual", "N Deteksi", "TP", "FP", "FN",
                    "Precision", "Recall", "F1", "Mean Dist (m)", "Median Dist (m)",
                    "RMSE Dist (m)", "Max Dist (m)", "Mode Matching", "Kelas Diperhitungkan?"]

    # Kumpulkan kolom atribut numerik (union dari semua model) supaya bisa dirata-ratakan
    # di sheet Ringkasan berdasarkan pasangan TP yang sudah di-join dengan atribut manual.
    infer_numeric_keys = []
    for r in model_results:
        for k in _numeric_attr_keys(r.get("attrs", [])):
            if k not in infer_numeric_keys:
                infer_numeric_keys.append(k)
    manual_numeric_keys = _numeric_attr_keys(manual_attrs)

    avg_headers = [f"Avg Infer.{k}" for k in infer_numeric_keys] + \
                  [f"Avg Manual.{k} (TP)" for k in manual_numeric_keys]
    headers = base_headers + avg_headers + ["Catatan"]
    ws.append(headers)

    # Model dengan F1 tertinggi di-highlight -- acuan cepat "mana yang terbaik".
    f1_list = [r["metrics"]["f1"] for r in model_results]
    best_idx = int(np.argmax(f1_list)) if f1_list else -1

    for i, r in enumerate(model_results):
        m = r["metrics"]
        infer_attrs = r.get("attrs", [])
        row = [
            r["name"], m["n_manual"], m["n_infer"], m["tp"], m["fp"], m["fn"],
            round(m["precision"], 4), round(m["recall"], 4), round(m["f1"], 4),
            round(m["mean_dist"], 4), round(m["median_dist"], 4),
            round(m["rmse_dist"], 4), round(m["max_dist"], 4),
            "Containment" if m.get("match_mode") == "containment" else "Jarak (Distance)",
            "Ya" if m.get("class_aware") else "Tidak (posisi saja)",
        ]
        # Rata-rata atribut infer dihitung dari SEMUA deteksi model ini (bukan cuma TP),
        # supaya tetap terisi walau jumlah TP sedikit/nol.
        for k in infer_numeric_keys:
            avg = _numeric_avg(a.get(k) for a in infer_attrs)
            row.append(round(avg, 4) if avg is not None else "")
        # Rata-rata atribut manual dihitung dari titik manual yang berhasil di-join (TP) saja,
        # karena atribut manual cuma bermakna untuk pasangan yang benar-benar match.
        for k in manual_numeric_keys:
            joined_vals = [manual_attrs[m_idx].get(k) for _, m_idx, _ in r["matches"]
                           if m_idx < len(manual_attrs)]
            avg = _numeric_avg(joined_vals)
            row.append(round(avg, 4) if avg is not None else "")
        row.append("\u2605 F1 Tertinggi" if i == best_idx else "")
        ws.append(row)

    last_row = ws.max_row
    _style_header_row(ws, n_cols=len(headers))
    _apply_number_formats(ws, headers, start_row=2)
    _apply_borders(ws, max_row=last_row, max_col=len(headers))

    best_row_abs = 2 + best_idx if best_idx >= 0 else None
    if best_row_abs:
        fill = PatternFill("solid", fgColor=_COLOR_BEST)
        for c in range(1, len(headers) + 1):
            ws.cell(row=best_row_abs, column=c).fill = fill

    _apply_row_banding(ws, header_row=1, last_row=last_row, n_cols=len(headers),
                       skip_fill={best_row_abs} if best_row_abs else None)
    _autofit_columns(ws)
    ws.freeze_panes = "A2"

    _write_glossary_sheet(wb, threshold)

    ws_info = wb.create_sheet("Info")
    ws_info.append(["Keterangan", "Nilai"])
    _style_header_row(ws_info, n_cols=2)
    ws_info.append(["Threshold jarak (meter)", threshold])
    ws_info.append(["Jumlah model dibandingkan", len(model_results)])
    ws_info.append(["Model dengan F1 tertinggi", model_results[best_idx]["name"] if best_idx >= 0 else "-"])
    ws_info.append(["Dibuat oleh", "Sawit Vision - Pembanding Model"])
    _apply_borders(ws_info)
    _autofit_columns(ws_info)
    ws_info.column_dimensions["A"].width = 28

    used_names = set()
    for r in model_results:
        sheet_name = _safe_sheet_name(r["name"], used_names)
        sh = wb.create_sheet(sheet_name)

        infer_attrs = r.get("attrs", [])
        # Nama kolom atribut infer & manual (union), di-prefix supaya jelas asalnya
        # dan tidak bentrok kalau ada nama kolom yang sama di kedua sumber data.
        infer_attr_keys = []
        for a in infer_attrs:
            for k in a.keys():
                if k not in infer_attr_keys:
                    infer_attr_keys.append(k)
        manual_attr_keys = []
        for a in manual_attrs:
            for k in a.keys():
                if k not in manual_attr_keys:
                    manual_attr_keys.append(k)

        header_row = (["status", "manual_idx", "infer_idx", "distance_m",
                        "infer_x", "infer_y", "manual_x", "manual_y"]
                      + [f"infer.{k}" for k in infer_attr_keys]
                      + [f"manual.{k}" for k in manual_attr_keys])
        sh.append(header_row)

        infer_xy = r["xy"]

        def _infer_attr_row(i_idx):
            a = infer_attrs[i_idx] if i_idx < len(infer_attrs) else {}
            return [a.get(k, "") for k in infer_attr_keys]

        def _manual_attr_row(m_idx):
            a = manual_attrs[m_idx] if m_idx < len(manual_attrs) else {}
            return [a.get(k, "") for k in manual_attr_keys]

        # TP: join langsung atribut manual + infer dalam satu baris, ini yang tadinya kosong.
        for i_idx, m_idx, d in r["matches"]:
            sh.append(["TP", m_idx, i_idx, round(d, 4),
                       float(infer_xy[i_idx][0]), float(infer_xy[i_idx][1]),
                       float(manual_xy[m_idx][0]), float(manual_xy[m_idx][1])]
                      + _infer_attr_row(i_idx)
                      + _manual_attr_row(m_idx))
        for i_idx in r["fp"]:
            sh.append(["FP", "", i_idx, "",
                       float(infer_xy[i_idx][0]), float(infer_xy[i_idx][1]), "", ""]
                      + _infer_attr_row(i_idx)
                      + [""] * len(manual_attr_keys))
        for m_idx in r["fn"]:
            sh.append(["FN", m_idx, "", "", "", "",
                       float(manual_xy[m_idx][0]), float(manual_xy[m_idx][1])]
                      + [""] * len(infer_attr_keys)
                      + _manual_attr_row(m_idx))

        n_cols = len(header_row)
        sh_last_row = sh.max_row
        _style_header_row(sh, n_cols=n_cols)
        _apply_number_formats(sh, header_row, start_row=2)
        _apply_borders(sh, max_row=sh_last_row, max_col=n_cols)

        # Warnai tiap baris sesuai status: TP hijau, FP merah, FN kuning --
        # supaya bisa langsung "dipindai mata" tanpa baca satu-satu.
        status_fill = {"TP": _COLOR_TP, "FP": _COLOR_FP, "FN": _COLOR_FN}
        for row_idx in range(2, sh_last_row + 1):
            status = sh.cell(row=row_idx, column=1).value
            color = status_fill.get(status)
            if color:
                fill = PatternFill("solid", fgColor=color)
                for c in range(1, n_cols + 1):
                    ws_cell = sh.cell(row=row_idx, column=c)
                    if ws_cell.fill.fgColor.rgb in (None, "00000000"):
                        ws_cell.fill = fill

        _autofit_columns(sh)
        sh.freeze_panes = "A2"

    wb.save(output_path)


# ============================================================
# EXPORT TITIK HASIL (TP / FP / FN) -- Shapefile + GeoJSON + CSV
# ============================================================
# Selain ringkasan Excel, tiap titik hasil pencocokan (benar/salah/terlewat)
# juga diekspor sebagai layer titik siap-pakai di GIS (QGIS dsb), supaya bisa
# langsung dilihat/diberi simbol beda per status di atas peta -- tidak cuma
# angka di tabel.

def _safe_filename(name: str) -> str:
    """Bikin nama file aman dari nama model bebas (hilangkan karakter yang
    tidak valid untuk nama file di Windows/Linux)."""
    keep = "".join(c if (c.isalnum() or c in ("_", "-")) else "_" for c in name.strip())
    keep = keep.strip("_")
    return keep or "model"


def _dbf_safe_fields(keys, prefix: str):
    """Nama kolom DBF (shapefile) maksimal 10 karakter. Fungsi ini membuat
    nama field pendek & unik (dengan prefix i_/m_ supaya jelas asalnya infer
    atau manual) dari nama kolom atribut asli yang bisa panjang/bebas."""
    used = set()
    mapping = {}
    for k in keys:
        base = (prefix + "".join(c for c in str(k) if c.isalnum()))[:10] or (prefix + "f")
        candidate = base
        n = 1
        while candidate in used:
            n += 1
            suffix = str(n)
            candidate = base[: 10 - len(suffix)] + suffix
        used.add(candidate)
        mapping[k] = candidate
    return mapping


def _build_status_rows(manual_xy, infer_xy, matches, fp, fn, infer_attrs, manual_attrs):
    """Satukan TP + FP + FN jadi satu daftar baris titik dengan kolom "status",
    supaya tiap model cukup 1 layer (bukan 3 file terpisah) dan tinggal
    difilter/diberi simbol beda per status di GIS."""
    rows = []
    for i_idx, m_idx, d in matches:
        rows.append({
            "status": "TP", "manual_idx": m_idx, "infer_idx": i_idx,
            "distance_m": round(float(d), 4),
            "x": float(infer_xy[i_idx][0]), "y": float(infer_xy[i_idx][1]),
            "_infer": infer_attrs[i_idx] if i_idx < len(infer_attrs) else {},
            "_manual": manual_attrs[m_idx] if m_idx < len(manual_attrs) else {},
        })
    for i_idx in fp:
        rows.append({
            "status": "FP", "manual_idx": None, "infer_idx": i_idx, "distance_m": None,
            "x": float(infer_xy[i_idx][0]), "y": float(infer_xy[i_idx][1]),
            "_infer": infer_attrs[i_idx] if i_idx < len(infer_attrs) else {},
            "_manual": {},
        })
    for m_idx in fn:
        rows.append({
            "status": "FN", "manual_idx": m_idx, "infer_idx": None, "distance_m": None,
            "x": float(manual_xy[m_idx][0]), "y": float(manual_xy[m_idx][1]),
            "_infer": {},
            "_manual": manual_attrs[m_idx] if m_idx < len(manual_attrs) else {},
        })
    return rows


def _read_crs_wkt(source_path):
    """Ambil definisi CRS dari file vektor apapun sebagai WKT string.
    - .shp    -> baca file .prj di sebelahnya (kalau ada)
    - .gpkg   -> baca dari tabel gpkg_spatial_ref_sys (kolom 'definition')
                 sesuai srs_id yang dipakai layer geometri pertama
    - .geojson/.json -> baca root "crs" (GeoJSON lama) kalau ada; kalau tidak
                 ada, GeoJSON RFC 7946 default = WGS84 (EPSG:4326) --
                 return WKT WGS84 sebagai default aman.

    Return string WKT, atau None kalau benar-benar tidak bisa ditentukan.

    Ini pengganti _copy_prj_if_available lama yang HANYA jalan untuk sumber
    .shp -- akibatnya kalau ground truth manual disimpan sebagai .gpkg,
    shapefile output tidak pernah dapat .prj dan titik-titiknya "lompat"
    keluar extent raster saat dibuka di QGIS.
    """
    if not source_path:
        return None
    ext = os.path.splitext(source_path)[1].lower()

    if ext == ".shp":
        prj = os.path.splitext(source_path)[0] + ".prj"
        if os.path.isfile(prj):
            try:
                with open(prj, "r", encoding="utf-8", errors="ignore") as f:
                    return f.read().strip() or None
            except OSError:
                return None
        return None

    if ext == ".gpkg":
        try:
            conn = sqlite3.connect(source_path)
            try:
                cur = conn.cursor()
                cur.execute("SELECT table_name, srs_id FROM gpkg_geometry_columns")
                row = cur.fetchone()
                if not row:
                    return None
                _, srs_id = row
                cur.execute(
                    "SELECT definition FROM gpkg_spatial_ref_sys WHERE srs_id = ?",
                    (srs_id,),
                )
                d = cur.fetchone()
                if d and d[0] and d[0].lower() != "undefined":
                    return d[0].strip()
            finally:
                conn.close()
        except (sqlite3.Error, OSError):
            return None
        return None

    if ext in (".geojson", ".json"):
        try:
            with open(source_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except (OSError, json.JSONDecodeError):
            return None
        crs = data.get("crs")
        if isinstance(crs, dict):
            props = crs.get("properties") or {}
            name = props.get("name") or ""
            # Contoh "urn:ogc:def:crs:EPSG::32749" -> ambil "32749"
            if "EPSG" in name.upper():
                digits = "".join(ch for ch in name if ch.isdigit())
                if digits:
                    wkt = _wkt_from_epsg(digits)
                    if wkt:
                        return wkt
            # crs eksplisit ada tapi kodenya tidak bisa di-resolve -- JANGAN
            # diam-diam anggap WGS84 (itu penyebab titik lompat keluar
            # raster kalau aslinya UTM). Biarkan None supaya caller coba
            # sumber CRS lain (mis. .prj milik hasil model).
            return None
        # RFC 7946: GeoJSON TANPA "crs" eksplisit = wajib WGS84 (lon/lat).
        return _WKT_WGS84

    return None


# WKT WGS84 (EPSG:4326) sebagai fallback aman untuk GeoJSON RFC 7946.
_WKT_WGS84 = (
    'GEOGCS["WGS 84",DATUM["WGS_1984",SPHEROID["WGS 84",6378137,298.257223563,'
    'AUTHORITY["EPSG","7030"]],AUTHORITY["EPSG","6326"]],'
    'PRIMEM["Greenwich",0,AUTHORITY["EPSG","8901"]],'
    'UNIT["degree",0.0174532925199433,AUTHORITY["EPSG","9122"]],'
    'AUTHORITY["EPSG","4326"]]'
)
try:
    from pyproj import CRS as _PyprojCRS
    from pyproj.exceptions import CRSError as _PyprojCRSError
    from pyproj.enums import WktVersion as _PyprojWktVersion
except ImportError:  # seharusnya tidak pernah kejadian di exe hasil build
    # (pyproj dibundel lewat build.spec), tapi dijaga biar app tetap jalan
    # walau lebih terbatas cakupan CRS-nya kalau entah kenapa hilang.
    _PyprojCRS = None


def _wkt_from_epsg(epsg_code):
    """Bangun WKT dari kode EPSG APAPUN lewat database EPSG resmi (pyproj) --
    UTM zona manapun, geografis, datum lokal, proyeksi apapun, semua otomatis
    benar tanpa perlu daftar manual per kamera/CRS baru.

    Kalau pyproj entah kenapa tidak tersedia di runtime, fallback ke builder
    manual yang cuma menangani WGS84 (4326) + UTM WGS84 zona 1-60 N/S --
    supaya app tetap jalan (cakupan CRS lebih sempit) daripada crash total.

    PENTING: return None kalau kode EPSG-nya benar-benar tidak bisa
    di-resolve -- JANGAN diam-diam anggap WGS84. Versi lama kode ini pakai
    dict statis {"4326": WGS84} dan fallback ke WGS84 untuk EPSG apapun yang
    tidak ada di dict (termasuk UTM seperti 32748) -- akibatnya file manual
    ground truth yang eksplisit UTM malah ditulis ulang dengan tag CRS WGS84,
    padahal koordinatnya tetap meter UTM. QGIS lalu membaca titik itu seolah
    lon/lat derajat -> titik "lompat" jauh keluar dari extent raster (paling
    kelihatan di FN karena koordinatnya murni dari file manual, tapi TP/FP
    dari model yang sama ikut kena juga).
    """
    try:
        code = int(epsg_code)
    except (TypeError, ValueError):
        return None

    if _PyprojCRS is not None:
        try:
            # WKT1_GDAL (bukan default WKT2) supaya formatnya tetap pakai
            # blok AUTHORITY["EPSG","..."] klasik -- kompatibel dengan
            # _extract_epsg_from_wkt() di bawah dan paling luas didukung
            # software GIS lama (ArcGIS, QGIS versi lawas, dst).
            return _PyprojCRS.from_epsg(code).to_wkt(_PyprojWktVersion.WKT1_GDAL)
        except _PyprojCRSError:
            return None  # kode EPSG tidak valid/tidak dikenal pyproj

    return _wkt_from_epsg_fallback(code)


def _wkt_from_epsg_fallback(code):
    """Fallback manual kalau pyproj tidak tersedia di runtime. Cuma cover
    WGS84 (4326) + UTM WGS84 zona 1-60 N/S (326xx/327xx) -- termasuk
    32748/32749/32750 yang umum dipakai untuk citra drone Indonesia."""
    if code == 4326:
        return _WKT_WGS84
    if 32601 <= code <= 32660:
        zone, hemi, false_northing = code - 32600, "N", 0
    elif 32701 <= code <= 32760:
        zone, hemi, false_northing = code - 32700, "S", 10000000
    else:
        return None
    central_meridian = -183 + zone * 6
    return (
        f'PROJCS["WGS 84 / UTM zone {zone}{hemi}",'
        'GEOGCS["WGS 84",DATUM["WGS_1984",'
        'SPHEROID["WGS 84",6378137,298.257223563,AUTHORITY["EPSG","7030"]],'
        'AUTHORITY["EPSG","6326"]],PRIMEM["Greenwich",0,'
        'AUTHORITY["EPSG","8901"]],UNIT["degree",0.0174532925199433,'
        'AUTHORITY["EPSG","9122"]],AUTHORITY["EPSG","4326"]],'
        'PROJECTION["Transverse_Mercator"],'
        'PARAMETER["latitude_of_origin",0],'
        f'PARAMETER["central_meridian",{central_meridian}],'
        'PARAMETER["scale_factor",0.9996],'
        'PARAMETER["false_easting",500000],'
        f'PARAMETER["false_northing",{false_northing}],'
        'UNIT["metre",1,AUTHORITY["EPSG","9001"]],'
        'AXIS["Easting",EAST],AXIS["Northing",NORTH],'
        f'AUTHORITY["EPSG","{code}"]]'
    )


def _write_prj(dest_shp_path, wkt):
    """Tulis file .prj di sebelah shapefile output. Aman dipanggil dengan
    wkt=None (dilewati diam-diam)."""
    if not wkt:
        return False
    try:
        prj_path = os.path.splitext(dest_shp_path)[0] + ".prj"
        with open(prj_path, "w", encoding="utf-8") as f:
            f.write(wkt)
        return True
    except OSError:
        return False


def _extract_epsg_from_wkt(wkt):
    """Ekstrak kode EPSG dari WKT string. WKT biasanya diakhiri:
    ...AUTHORITY["EPSG","32750"]]. Kita ambil angka di dalam AUTHORITY
    "EPSG",... yang paling akhir (paling luar) -- itu EPSG asli si CRS,
    bukan sub-komponen seperti datum atau ellipsoid.

    Return string EPSG (mis. "32750") atau None kalau tidak ketemu.

    Ini dipakai untuk isi property "crs" di GeoJSON output supaya QGIS
    tidak salah anggap sebagai default WGS84 (EPSG:4326) padahal koordinatnya
    UTM.
    """
    if not wkt:
        return None
    import re
    # Cari SEMUA occurrence AUTHORITY["EPSG","<digits>"] lalu ambil yang terakhir
    # -- yang terakhir di WKT well-formed selalu EPSG dari CRS terluar.
    matches = re.findall(r'AUTHORITY\s*\[\s*"EPSG"\s*,\s*"(\d+)"\s*\]', wkt)
    return matches[-1] if matches else None


def _write_points_shapefile(path, rows, infer_keys, manual_keys):
    infer_map = _dbf_safe_fields(infer_keys, "i_")
    manual_map = _dbf_safe_fields(manual_keys, "m_")

    with shapefile.Writer(path, shapeType=shapefile.POINT) as shp:
        shp.field("status", "C", size=8)
        shp.field("manual_idx", "N", size=10)
        shp.field("infer_idx", "N", size=10)
        shp.field("dist_m", "N", size=12, decimal=4)
        for k in infer_keys:
            shp.field(infer_map[k], "C", size=50)
        for k in manual_keys:
            shp.field(manual_map[k], "C", size=50)

        for row in rows:
            shp.point(row["x"], row["y"])
            record = [row["status"], row["manual_idx"], row["infer_idx"], row["distance_m"]]
            for k in infer_keys:
                v = row["_infer"].get(k)
                record.append("" if v is None else str(v))
            for k in manual_keys:
                v = row["_manual"].get(k)
                record.append("" if v is None else str(v))
            shp.record(*record)


def _write_points_geojson(path, rows, infer_keys, manual_keys, epsg=None):
    """Tulis titik ke .geojson. Kalau `epsg` diberikan, sisipkan property
    "crs" di root FeatureCollection supaya QGIS dan software GIS lain
    langsung tahu koordinatnya di CRS apa.

    Kenapa ini penting: RFC 7946 (GeoJSON modern) bilang kalau tidak ada
    property "crs", default = WGS84 (EPSG:4326). Padahal kita nulis
    koordinat UTM. Kalau tidak explicit tulis "crs", QGIS akan render
    titik UTM sebagai lat/lon --> keluar extent raster jauh.

    Format "crs" yang dipakai: OGC URN format (yang paling kompatibel
    dengan QGIS, GDAL, ArcGIS, dsb):
        {"type": "name", "properties": {"name": "urn:ogc:def:crs:EPSG::<code>"}}
    """
    features = []
    for row in rows:
        props = {
            "status": row["status"], "manual_idx": row["manual_idx"],
            "infer_idx": row["infer_idx"], "distance_m": row["distance_m"],
        }
        for k in infer_keys:
            props[f"infer.{k}"] = row["_infer"].get(k)
        for k in manual_keys:
            props[f"manual.{k}"] = row["_manual"].get(k)
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [row["x"], row["y"]]},
            "properties": props,
        })
    fc = {"type": "FeatureCollection", "features": features}
    if epsg:
        fc["crs"] = {
            "type": "name",
            "properties": {"name": f"urn:ogc:def:crs:EPSG::{epsg}"},
        }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(fc, f, indent=2, default=str)


def _write_points_csv(path, rows, infer_keys, manual_keys):
    import csv as _csv
    fieldnames = (["status", "manual_idx", "infer_idx", "distance_m", "x", "y"]
                  + [f"infer.{k}" for k in infer_keys] + [f"manual.{k}" for k in manual_keys])
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = _csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            out = {"status": row["status"], "manual_idx": row["manual_idx"],
                   "infer_idx": row["infer_idx"], "distance_m": row["distance_m"],
                   "x": row["x"], "y": row["y"]}
            for k in infer_keys:
                out[f"infer.{k}"] = row["_infer"].get(k, "")
            for k in manual_keys:
                out[f"manual.{k}"] = row["_manual"].get(k, "")
            writer.writerow(out)


def export_comparison_points(output_dir, manual_xy, model_results, manual_path=None, manual_attrs=None):
    """
    Ekspor titik hasil pencocokan tiap model sebagai Shapefile + GeoJSON + CSV
    ke dalam output_dir, dalam 2 bentuk:
      1. File TERPISAH per status: "<model>_TP.*", "<model>_FP.*", "<model>_FN.*"
         -- masing-masing cuma berisi titik dengan status itu saja. Ini yang
         paling gampang dipakai kalau mau kasih simbol/warna beda per layer
         langsung di QGIS tanpa perlu filter dulu. Status yang jumlah titiknya
         nol dilewati (tidak bikin file kosong).
      2. Satu file GABUNGAN "<model>_hasil.*" berisi SEMUA titik (TP+FP+FN)
         dengan kolom "status" -- buat yang lebih suka 1 layer + filter/style
         berdasarkan atribut.

    model_results: sama seperti pada export_comparison_excel (list of dict
        dengan "name", "xy", "matches", "fp", "fn", "attrs").
    manual_path: path file centroid manual asli (dipakai buat coba salin
        .prj-nya kalau sumbernya .shp, supaya CRS ikut terbawa).

    Return: list of dict per model:
        {"name": str,
         "combined": {"shp":.., "geojson":.., "csv":..},
         "by_status": {"TP": {...} atau None, "FP": {...} atau None, "FN": {...} atau None}}
    """
    manual_attrs = manual_attrs or []
    os.makedirs(output_dir, exist_ok=True)

    # Tentukan CRS output SEKALI di awal, prioritas:
    #   1. CRS dari file manual (ground truth) -- ini yang paling benar karena
    #      koordinat FN memang datang dari sini.
    #   2. Fallback: CRS dari file inference pertama yang punya .prj (biasanya
    #      hasil Sawit Vision selalu punya, karena save_shapefile di
    #      inference_core.py menulisnya dari raster).
    # Kalau dua-duanya tidak ada, .prj tidak ditulis (perilaku lama), dan
    # QGIS akan minta CRS manual saat file dibuka.
    output_wkt = _read_crs_wkt(manual_path)
    if not output_wkt:
        for r in model_results:
            candidate_wkt = _read_crs_wkt(r.get("path"))
            if candidate_wkt:
                output_wkt = candidate_wkt
                break

    # Ekstrak EPSG code dari WKT -- dipakai untuk isi property "crs" di
    # GeoJSON output supaya QGIS/GDAL langsung mengenali CRS-nya, bukan
    # fallback ke WGS84 (yang bikin titik UTM lompat keluar bumi).
    output_epsg = _extract_epsg_from_wkt(output_wkt)

    outputs = []
    used_names = set()
    for r in model_results:
        rows = _build_status_rows(
            manual_xy, r["xy"], r["matches"], r["fp"], r["fn"],
            r.get("attrs", []), manual_attrs,
        )

        infer_keys = []
        manual_keys = []
        for row in rows:
            for k in row["_infer"]:
                if k not in infer_keys:
                    infer_keys.append(k)
            for k in row["_manual"]:
                if k not in manual_keys:
                    manual_keys.append(k)

        base_name = _safe_filename(r["name"])
        candidate = base_name
        n = 1
        while candidate in used_names:
            n += 1
            candidate = f"{base_name}_{n}"
        used_names.add(candidate)

        # Setiap model dapat sub-folder sendiri, supaya kalau bandingin
        # beberapa model (mis. Multispectral vs RGB) file-nya tidak
        # numpuk campur jadi satu. Di dalam sub-folder model, TP/FP/FN/
        # Gabungan masing-masing juga punya sub-foldernya sendiri.
        model_dir = os.path.join(output_dir, candidate)
        os.makedirs(model_dir, exist_ok=True)

        def _write_set(subdir_name, suffix, subset_rows):
            target_dir = os.path.join(model_dir, subdir_name)
            os.makedirs(target_dir, exist_ok=True)
            shp_path = os.path.join(target_dir, f"{candidate}_{suffix}.shp")
            geojson_path = os.path.join(target_dir, f"{candidate}_{suffix}.geojson")
            csv_path = os.path.join(target_dir, f"{candidate}_{suffix}.csv")
            _write_points_shapefile(shp_path, subset_rows, infer_keys, manual_keys)
            _write_points_geojson(geojson_path, subset_rows, infer_keys, manual_keys,
                                   epsg=output_epsg)
            _write_points_csv(csv_path, subset_rows, infer_keys, manual_keys)
            _write_prj(shp_path, output_wkt)
            return {"shp": shp_path, "geojson": geojson_path, "csv": csv_path}

        by_status = {}
        for status in ("TP", "FP", "FN"):
            subset = [row for row in rows if row["status"] == status]
            by_status[status] = _write_set(status, status, subset) if subset else None

        combined = _write_set("Gabungan", "hasil", rows)

        outputs.append({
            "name": r["name"], "folder": model_dir,
            "combined": combined, "by_status": by_status,
        })

    return outputs
